import csv
import io
import unittest

import pandas as pd
from openpyxl import load_workbook

from app import _dataframe_to_csv_bytes, _dataframe_to_xlsx_bytes


class ExportSecurityTests(unittest.TestCase):
    def setUp(self) -> None:
        self.values = [
            "=SUM(1,1)",
            "+cmd",
            "-1+1",
            "@SUM(1,1)",
            "\t=SUM(1,1)",
            "\r=SUM(1,1)",
            "\n=SUM(1,1)",
            "ordinary text",
        ]
        self.frame = pd.DataFrame({"value": self.values, "number": range(len(self.values))})

    def test_csv_escapes_formula_prefixes_without_mutating_source(self) -> None:
        rows = list(
            csv.DictReader(
                io.StringIO(_dataframe_to_csv_bytes(self.frame).decode("utf-8"))
            )
        )

        self.assertEqual([row["value"] for row in rows[:7]], [f"'{v}" for v in self.values[:7]])
        self.assertEqual(rows[7]["value"], "ordinary text")
        self.assertEqual(self.frame["value"].tolist(), self.values)

    def test_xlsx_escapes_formula_prefixes_and_preserves_numbers(self) -> None:
        workbook = load_workbook(io.BytesIO(_dataframe_to_xlsx_bytes(self.frame)))
        worksheet = workbook["rosters"]
        exported_values = [worksheet.cell(row=row, column=1).value for row in range(2, 10)]
        exported_numbers = [worksheet.cell(row=row, column=2).value for row in range(2, 10)]

        self.assertEqual(exported_values[:7], [f"'{v}" for v in self.values[:7]])
        self.assertEqual(exported_values[7], "ordinary text")
        self.assertEqual(exported_numbers, list(range(8)))


if __name__ == "__main__":
    unittest.main()
